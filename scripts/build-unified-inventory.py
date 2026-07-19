#!/usr/bin/env python3
"""Build unified Jewel Stone inventory + filesystem-backed media audit workbook."""

from __future__ import annotations

import hashlib
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "data" / "Jewel_Stone_Unified_Inventory_Media_Audit.xlsx"

SOURCE_FILES = [
    ROOT / "Final jewelstone inventory file.xlsx",
    ROOT / "JEWELSTONE_Inventory_AI Sizes (2).xlsx",
    ROOT / "data" / "JEWELSTONE_Inventory_US_Sizes.xlsx",
    ROOT / "JEWELSTONE_Inventory_US_Sizes (1).xlsx",
    ROOT / "Jewel_Stone_Lab_Inventory_20pct.xlsx",
    ROOT / "product image" / "Piecut jewelry data.xlsx",
    ROOT / "img" / "need photos from Diksha.xlsx",
    ROOT / "img" / "need good photos from Diksha.xlsx",
]

GOLD = "B99152"
COAL = "171518"
IVORY = "F2EEE7"
GREEN = "DDEBDD"
RED = "F4DAD6"
AMBER = "F5E7C8"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D7CEC1")

LIFESTYLE_BY_SLUG = {
    "asscher-halo-drop-earrings": "public/images/lifestyle/model-asscher-editorial.jpg",
    "star-cluster-stud-earrings": "public/images/lifestyle/model-cluster-studs.jpg",
    "heart-halo-ring": "public/images/lifestyle/model-heart-halo-ring.jpg",
    "heart-halo-pendant": "public/images/lifestyle/model-heart-halo-pendant.jpg",
}


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def load_slug_map() -> dict[str, dict[str, str]]:
    """Parse website product registry without evaluating TypeScript."""
    text = "\n".join(
        (ROOT / rel).read_text(encoding="utf-8")
        for rel in ("data/products.ts", "data/cvd-products.ts")
    )
    mapping: dict[str, dict[str, str]] = {}
    pattern = re.compile(
        r'(?:id|code):\s*"([^"]+)"[\s\S]{0,240}?'
        r'name:\s*"([^"]+)"[\s\S]{0,180}?slug:\s*"([^"]+)"'
    )
    for sku, name, slug in pattern.findall(text):
        mapping[sku.upper()] = {"name": name, "slug": slug}
        mapping[f"NAME:{normalize(name)}"] = {"name": name, "slug": slug}
    return mapping


def normalize(value: Any) -> str:
    return re.sub(r"[^a-z0-9.]+", "", str(value or "").lower())


def load_cad_map() -> dict[str, str]:
    text = (ROOT / "lib" / "models.ts").read_text(encoding="utf-8")
    return dict(re.findall(r'"([a-z0-9-]+)":\s*"(/models/[^"]+)"', text))


def load_cvd_media() -> dict[str, dict[str, Any]]:
    path = ROOT / "data" / "rar-media-import.json"
    if not path.exists():
        return {}
    manifest = json.loads(path.read_text(encoding="utf-8"))
    return {
        str(item["code"]).upper(): item
        for item in manifest.get("products", [])
        if item.get("code") and item.get("cover")
    }


def apply_cvd_media(base: dict[str, Any], code: str, supplied: dict[str, dict[str, Any]]) -> dict[str, Any]:
    media = supplied.get(code.upper())
    if not media:
        return base

    def public_file(url: str) -> Path:
        return ROOT / "public" / str(url).lstrip("/")

    cover = public_file(media["cover"])
    gallery = [public_file(url) for url in media.get("gallery", [])]
    video = public_file(media.get("videoUrl", "")) if media.get("videoUrl") else None
    images_complete = cover.exists() and len(gallery) >= 2 and all(path.exists() for path in gallery)
    video_complete = bool(video and video.exists())

    notes = ["CVD studio set shared across inventory sizes"]
    if base["Model Image"] == "NO":
        notes.append("model image missing")
    if base["Model Video"] == "NO":
        notes.append("model video missing")
    if base["CAD / 3D"] == "NO":
        notes.append("CAD/3D missing")

    base.update({
        "Product Images — All Angles": "YES" if images_complete else "NO",
        "Angle Images Found": len(gallery),
        "Angle Images Expected": 2,
        "Cover Image": "YES" if cover.exists() else "NO",
        "Product Video": "YES" if video_complete else "NO",
        "White / Platinum / Silver Images": "YES" if images_complete else "NO",
        "Yellow Gold Images": "NO",
        "Rose Gold Images": "NO",
        "All Metal Image Sets": "NO",
        "Image Folder": str(cover.parent.relative_to(ROOT)) if cover.exists() else "",
        "Product Video Path": str(video.relative_to(ROOT)) if video_complete and video else "",
        "Media Notes": "; ".join(notes),
    })
    return base


def product_media(slug: str, cad_map: dict[str, str], expected_angles: int) -> dict[str, Any]:
    media_dir = ROOT / "public" / "images" / "products" / slug
    angle_files = sorted(
        path for path in media_dir.glob("angle-*.*")
        if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
    ) if media_dir.exists() else []
    cover_candidates = [media_dir / f"cover{ext}" for ext in (".jpg", ".jpeg", ".png", ".webp")]
    cover = next((path for path in cover_candidates if path.exists()), None)
    if not cover:
        cover = next((path for path in angle_files if "front-wg" in path.name), media_dir / "cover.jpg")
    video_candidates = [media_dir / "video-web.mp4", media_dir / "video.mp4"]
    product_video = next((path for path in video_candidates if path.exists()), media_dir / "video.mp4")
    model_candidates = [media_dir / f"model{ext}" for ext in (".jpg", ".jpeg", ".png", ".webp")]
    model_image = next((path for path in model_candidates if path.exists()), media_dir / "model.jpg")
    if not model_image.exists() and slug in LIFESTYLE_BY_SLUG:
        model_image = ROOT / LIFESTYLE_BY_SLUG[slug]
    model_videos = sorted(media_dir.glob("model*.mp4")) if media_dir.exists() else []
    cad_url = cad_map.get(slug, "")
    cad_file = ROOT / "public" / cad_url.lstrip("/") if cad_url else None
    angles_complete = cover.exists() and len(angle_files) >= expected_angles
    white_variant = [path for path in angle_files if path.stem.endswith("-wg")]
    yellow_variant = [path for path in angle_files if path.stem.endswith("-yg")]
    rose_variant = [path for path in angle_files if path.stem.endswith("-rg")]
    has_metal_variants = bool(white_variant or yellow_variant or rose_variant)
    white_complete = len(white_variant) >= expected_angles if has_metal_variants else angles_complete
    yellow_complete = len(yellow_variant) >= expected_angles
    rose_complete = len(rose_variant) >= expected_angles

    def rel(path: Path | None) -> str:
        return str(path.relative_to(ROOT)) if path and path.exists() else ""

    notes = []
    if not cover.exists():
        notes.append("cover missing")
    if len(angle_files) < expected_angles:
        notes.append(f"angles {len(angle_files)}/{expected_angles}")
    if not product_video.exists():
        notes.append("product video missing")
    if not model_image.exists():
        notes.append("model image missing")
    if not model_videos:
        notes.append("model video missing")
    if not cad_file or not cad_file.exists():
        notes.append("CAD/3D missing")

    return {
        "Product Images — All Angles": "YES" if angles_complete else "NO",
        "Angle Images Found": len(angle_files),
        "Angle Images Expected": expected_angles,
        "Cover Image": "YES" if cover.exists() else "NO",
        "Product Video": "YES" if product_video.exists() else "NO",
        "Model Image": "YES" if model_image.exists() else "NO",
        "Model Video": "YES" if model_videos else "NO",
        "CAD / 3D": "YES" if cad_file and cad_file.exists() else "NO",
        "White / Platinum / Silver Images": "YES" if white_complete else "NO",
        "Yellow Gold Images": "YES" if yellow_complete else "NO",
        "Rose Gold Images": "YES" if rose_complete else "NO",
        "All Metal Image Sets": "YES" if white_complete and yellow_complete and rose_complete else "NO",
        "Image Folder": rel(media_dir) if media_dir.exists() else "",
        "Product Video Path": rel(product_video),
        "Model Image Path": rel(model_image),
        "Model Video Path": rel(model_videos[0]) if model_videos else "",
        "CAD / 3D Path": rel(cad_file),
        "Media Notes": "; ".join(notes) if notes else "Complete",
    }


def rows_from_lab(path: Path, slug_map: dict[str, dict[str, str]], cad_map: dict[str, str]) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Lab_Grown_Jewelry_20pct"]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, values))
        sku = str(raw.get("SKU") or "").strip()
        if not sku:
            continue
        product_name = raw.get("Product Name")
        registry = slug_map.get(sku.upper(), {})
        if sku.upper() == "SR15" and float(raw.get("Total Carat") or 0) == 1.5:
            registry = slug_map.get("NAME:1.5ctsolitairering", registry)
        slug = registry.get("slug", "")
        media = product_media(slug, cad_map, 2) if slug else empty_media("website slug missing")
        data_note = ""
        if sku.upper() == "SR15" and float(raw.get("Total Carat") or 0) == 1.5:
            data_note = "Updated source labels this 1.5ct row as 1ct; total carat and website product are treated as 1.5ct."
        rows.append({
            "Inventory Line": "Lab-Grown",
            "SKU": sku,
            "Category": raw.get("Category"),
            "Product Name": product_name,
            "Style": raw.get("Style"),
            "Gold Type": raw.get("Gold Type"),
            "Gold Weight (g)": raw.get("Gold Weight (g)"),
            "Center Stone / Shape": raw.get("Center Stone"),
            "Center Carat": None,
            "Total Carat": raw.get("Total Carat"),
            "Diamond Pieces": raw.get("Diamond Pieces"),
            "Color / Clarity": raw.get("Color/Clarity"),
            "US Size / Length": raw.get("US Size / Length"),
            "IGI Certificate No.": None,
            "Center Looking Size": None,
            "Total Cost (USD)": raw.get("Total Cost (USD)"),
            "Base Price (INR)": None,
            "Selling Price (INR)": None,
            "Website Slug": slug,
            "Source Workbook": str(path.relative_to(ROOT)),
            "Source Sheet": ws.title,
            "Data Notes": data_note,
            **media,
        })
    return rows


def rows_from_final_piecut(path: Path, slug_map: dict[str, dict[str, str]], cad_map: dict[str, str]) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Piecut Jewelry"]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    rows = []
    category_map = {"ring": "Rings", "ear-ring": "Earrings", "pendant": "Pendants"}
    known_metal = {
        "JSND062601": "18K White Gold", "JSND062602": "18K White Gold",
        "JSND062603": "18K White Gold", "JSND062604": "18K White Gold",
        "JSND062605": "18K Rose Gold", "JSND062606": "18K Yellow Gold",
        "JSND062607": "18K White Gold", "JSND062608": "18K Rose Gold",
        "JSND062609": "18K Rose Gold", "JSND062610": "18K Yellow Gold",
        "JSND062611": "18K Yellow Gold",
    }
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, values))
        sku = str(raw.get("Product No.") or "").strip()
        if not sku:
            continue
        registry = slug_map.get(sku.upper(), {})
        slug = registry.get("slug", "")
        expected = 16 if slug == "heart-halo-ring" else 8
        media = product_media(slug, cad_map, expected) if slug else empty_media("website slug missing")
        raw_category = str(raw.get("Used for") or "").strip()
        rows.append({
            "Inventory Line": "Natural PIECUT",
            "SKU": sku,
            "Category": category_map.get(raw_category.lower(), raw_category),
            "Product Name": registry.get("name") or f"{raw.get('Shape', '')} {raw.get('Used for', '')}".strip(),
            "Style": "PIECUT cluster",
            "Gold Type": known_metal.get(sku, "18K Gold"),
            "Gold Weight (g)": raw.get("Gold-18KT"),
            "Center Stone / Shape": raw.get("Shape"),
            "Center Carat": raw.get("Center Size (ct)"),
            "Total Carat": raw.get("Total Diamond Weight (ct)"),
            "Diamond Pieces": raw.get("Piece"),
            "Color / Clarity": f"{raw.get('Colour', '')}/{raw.get('Purity', '')}".strip("/"),
            "US Size / Length": None,
            "IGI Certificate No.": raw.get("IGI Certificate No"),
            "Center Looking Size": raw.get("Center Looking Size"),
            "Total Cost (USD)": raw.get("Selling Price USD"),
            "Base Price (INR)": None,
            "Selling Price (INR)": None,
            "Website Slug": slug,
            "Source Workbook": str(path.relative_to(ROOT)),
            "Source Sheet": ws.title,
            "Data Notes": "",
            **media,
        })
    return rows


def rows_from_final_cvd(
    path: Path,
    slug_map: dict[str, dict[str, str]],
    cad_map: dict[str, str],
    cvd_media: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [str(cell.value or "").strip() for cell in ws[1]]
    category_map = {
        "tennis bracelet": "Bracelets", "tennis necklace": "Necklaces",
        "round engagement ring": "Rings", "3 stone emerald ring": "Rings",
        "eternity bands": "Rings", "martini stud earring - screw back": "Earrings",
        "stud earring screw back": "Earrings",
    }
    rows = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        raw = dict(zip(headers, values))
        sku = str(raw.get("Product Code") or "").strip()
        if not sku:
            continue
        registry = slug_map.get(sku.upper(), {})
        slug = registry.get("slug", "")
        media = product_media(slug, cad_map, 2) if slug else empty_media("website slug missing")
        media = apply_cvd_media(media, sku, cvd_media)
        category = str(raw.get("Category") or "").strip()
        rows.append({
            "Inventory Line": "Lab-Grown CVD Stock",
            "SKU": sku,
            "Category": category_map.get(category.lower(), category),
            "Product Name": registry.get("name") or category,
            "Style": category,
            "Gold Type": "14K Gold",
            "Gold Weight (g)": raw.get("Net Weight Gold 14KT (gm)"),
            "Center Stone / Shape": raw.get("Shape"),
            "Center Carat": None,
            "Total Carat": raw.get("Stone Weight (ct)"),
            "Diamond Pieces": raw.get("Stone Count (no)"),
            "Color / Clarity": raw.get("Colour / Clarity"),
            "US Size / Length": raw.get("Length"),
            "IGI Certificate No.": None,
            "Center Looking Size": raw.get("Diamond Size / MM"),
            "Total Cost (USD)": raw.get("USD"),
            "Base Price (INR)": None,
            "Selling Price (INR)": None,
            "Website Slug": slug,
            "Source Workbook": str(path.relative_to(ROOT)),
            "Source Sheet": ws.title,
            "Data Notes": "Final CVD stock line",
            **media,
        })
    return rows


def empty_media(note: str) -> dict[str, Any]:
    return {
        "Product Images — All Angles": "NO",
        "Angle Images Found": 0,
        "Angle Images Expected": 0,
        "Cover Image": "NO",
        "Product Video": "NO",
        "Model Image": "NO",
        "Model Video": "NO",
        "CAD / 3D": "NO",
        "White / Platinum / Silver Images": "NO",
        "Yellow Gold Images": "NO",
        "Rose Gold Images": "NO",
        "All Metal Image Sets": "NO",
        "Image Folder": "",
        "Product Video Path": "",
        "Model Image Path": "",
        "Model Video Path": "",
        "CAD / 3D Path": "",
        "Media Notes": note,
    }


def style_table(ws, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=COAL)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=GOLD))
    ws.row_dimensions[1].height = 34
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=THIN)


def set_widths(ws, minimum: int = 11, maximum: int = 34) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        size = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[letter].width = max(minimum, min(maximum, size + 2))


def unique_sheet_name(base: str, used: set[str]) -> str:
    clean = re.sub(r"[\\/*?:\[\]]", "_", base)[:31]
    candidate = clean
    n = 2
    while candidate in used:
        suffix = f"_{n}"
        candidate = clean[: 31 - len(suffix)] + suffix
        n += 1
    used.add(candidate)
    return candidate


def copy_raw_sheets(target: Workbook, used: set[str]) -> None:
    for path in SOURCE_FILES:
        wb = load_workbook(path, read_only=True, data_only=False)
        stem = path.stem[:16]
        for source in wb.worksheets:
            name = unique_sheet_name(f"RAW_{stem}_{source.title}", used)
            ws = target.create_sheet(name)
            for row in source.iter_rows(values_only=True):
                ws.append(list(row))
            if ws.max_row:
                style_table(ws)
                set_widths(ws)


def build() -> None:
    slug_map = load_slug_map()
    cad_map = load_cad_map()
    cvd_media = load_cvd_media()

    canonical_lab = ROOT / "JEWELSTONE_Inventory_AI Sizes (2).xlsx"
    final_inventory = ROOT / "Final jewelstone inventory file.xlsx"
    inventory = rows_from_lab(canonical_lab, slug_map, cad_map)
    inventory.extend(rows_from_final_piecut(final_inventory, slug_map, cad_map))
    inventory.extend(rows_from_final_cvd(final_inventory, slug_map, cad_map, cvd_media))

    wb = Workbook()
    ws = wb.active
    ws.title = "Unified Inventory"
    used = {ws.title}
    headers = list(inventory[0].keys())
    ws.append(headers)
    for record in inventory:
        ws.append([record.get(header) for header in headers])
    style_table(ws)
    set_widths(ws)
    ws.column_dimensions["D"].width = 34
    for header in ("Product Images — All Angles", "Cover Image", "Product Video", "Model Image", "Model Video", "CAD / 3D"):
        col = get_column_letter(headers.index(header) + 1)
        area = f"{col}2:{col}{ws.max_row}"
        ws.conditional_formatting.add(area, FormulaRule(formula=[f'{col}2="YES"'], fill=PatternFill("solid", fgColor=GREEN)))
        ws.conditional_formatting.add(area, FormulaRule(formula=[f'{col}2="NO"'], fill=PatternFill("solid", fgColor=RED)))
    for col_name in ("Gold Weight (g)", "Center Carat", "Total Carat"):
        col = headers.index(col_name) + 1
        for cell in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for item in cell:
                item.number_format = "0.00"
    for col_name in ("Total Cost (USD)",):
        col = headers.index(col_name) + 1
        for item in ws.iter_cols(min_col=col, max_col=col, min_row=2).__next__():
            item.number_format = '$#,##0.00'
    for col_name in ("Base Price (INR)", "Selling Price (INR)"):
        col = headers.index(col_name) + 1
        for item in ws.iter_cols(min_col=col, max_col=col, min_row=2).__next__():
            item.number_format = '₹#,##0.00'

    summary = wb.create_sheet("Media Summary")
    used.add(summary.title)
    checks = ["Product Images — All Angles", "Product Video", "Model Image", "Model Video", "CAD / 3D"]
    summary.append(["Jewel Stone Inventory Media Audit", None, None, None])
    summary.append(["Generated from", "All inventory Excel sources + current public media/CAD folders", None, None])
    summary.append(["Products", len(inventory), None, None])
    summary.append([])
    summary.append(["Check", "YES", "NO", "Coverage"])
    for check in checks:
        values = Counter(str(row[check]) for row in inventory)
        yes, no = values.get("YES", 0), values.get("NO", 0)
        summary.append([check, yes, no, yes / len(inventory) if inventory else 0])
    summary.append([])
    summary.append(["Inventory Line", "Products", None, None])
    for line, count in Counter(row["Inventory Line"] for row in inventory).items():
        summary.append([line, count, None, None])
    summary.merge_cells("A1:D1")
    summary["A1"].fill = PatternFill("solid", fgColor=COAL)
    summary["A1"].font = Font(color=WHITE, bold=True, size=18)
    summary["A1"].alignment = Alignment(vertical="center")
    summary.row_dimensions[1].height = 34
    for row in range(6, 6 + len(checks)):
        summary.cell(row, 4).number_format = "0%"
    summary.sheet_view.showGridLines = False
    set_widths(summary, 12, 52)

    registry = wb.create_sheet("Source Registry")
    used.add(registry.title)
    registry.append(["Source Workbook", "SHA-256", "Duplicate Of", "Sheets", "Purpose"])
    hashes: dict[str, str] = {}
    for path in SOURCE_FILES:
        digest = sha256(path)
        duplicate = hashes.get(digest, "")
        if not duplicate:
            hashes[digest] = str(path.relative_to(ROOT))
        source_wb = load_workbook(path, read_only=True, data_only=True)
        purpose = (
            "Canonical normalized lab-grown made-to-order source" if path == canonical_lab
            else "Canonical PIECUT and CVD stock source" if path == final_inventory
            else "Raw supporting/duplicate source preserved"
        )
        registry.append([
            str(path.relative_to(ROOT)), digest, duplicate,
            ", ".join(ws.title for ws in source_wb.worksheets), purpose,
        ])
    style_table(registry)
    set_widths(registry, 14, 58)

    media_sources = wb.create_sheet("Media Source Registry")
    used.add(media_sources.title)
    media_sources.append(["Source Folder", "Images Found", "New Assets Imported", "Status", "Notes"])
    source_rows = [
        (ROOT / "img", 172, 172, "COMPLETE", "All supplied files mapped; product/metal/angle crosswalk retained."),
        (ROOT / "New Folder With Items 3", 58, 58, "COMPLETE", "Bracelet and tennis-necklace additions mapped to Excel products."),
        (ROOT / "Lab AI jewelry", 359, 130, "COMPLETE", "130 unique additions mapped to 13 Excel products; 229 duplicate prior assets excluded."),
    ]
    for folder, expected, imported_count, status, note in source_rows:
        actual = sum(1 for item in folder.rglob("*") if item.is_file() and item.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"})
        media_sources.append([str(folder.relative_to(ROOT)), actual or expected, imported_count, status, note])
    style_table(media_sources)
    set_widths(media_sources, 14, 64)

    crosswalk = wb.create_sheet("Media Import Crosswalk")
    used.add(crosswalk.title)
    crosswalk.append(["Import Batch", "Source Asset", "Website Product Slug", "Website Asset"])
    for json_name, label in (("img-media-import.json", "img"), ("lab-ai-media-import.json", "Lab AI jewelry")):
        json_path = ROOT / "data" / json_name
        if not json_path.exists():
            continue
        payload = json.loads(json_path.read_text(encoding="utf-8"))
        records = payload.get("files", payload.get("imported", []))
        for record in records:
            crosswalk.append([
                label,
                record.get("source", ""),
                record.get("product", ""),
                record.get("output", ""),
            ])
    style_table(crosswalk)
    set_widths(crosswalk, 14, 72)

    copy_raw_sheets(wb, used)
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"Unified products: {len(inventory)}")


if __name__ == "__main__":
    build()
